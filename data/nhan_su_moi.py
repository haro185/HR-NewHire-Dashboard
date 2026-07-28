import os
import re
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==== CẤU HÌNH ====
EMAIL = "hao.tx@thanhhamedical.com"
API_KEY = "aPYs5c7EdyskN3IqBTm5iXSb9s27ufib"
SITE = "https://chat.mitalab.com"
STREAM = "[HR] THÔNG BÁO NHÂN SỰ MỚI"
TOPIC = "Thông báo nhân viên mới"
OUTPUT_FILE = "nhan_su_moi.xlsx"

FONT_NAME = "Arial"


# ---------------------------------------------------------------------------
# 1. LẤY TOÀN BỘ TIN NHẮN TỪ ZULIP
# ---------------------------------------------------------------------------
def fetch_messages():
    import zulip  # import cục bộ để phần còn lại của module test được mà không cần cài zulip
    client = zulip.Client(email=EMAIL, api_key=API_KEY, site=SITE)
    narrow = [
        {"operator": "stream", "operand": STREAM},
        {"operator": "topic", "operand": TOPIC},
    ]
    messages = []
    anchor = "oldest"
    while True:
        result = client.get_messages({
            "anchor": anchor,
            "num_before": 0,
            "num_after": 1000,
            "narrow": narrow,
        })
        if result.get("result") != "success":
            raise RuntimeError(f"Lỗi gọi API Zulip: {result}")
        batch = result["messages"]
        if not batch:
            break
        messages.extend(batch)
        if len(batch) < 1000:
            break
        anchor = batch[-1]["id"]
    return messages


# ---------------------------------------------------------------------------
# 2. PARSE BẢNG TRONG TỪNG TIN NHẮN
# ---------------------------------------------------------------------------
def parse_messages(messages):
    headers = None
    rows = []
    for m in messages:
        soup = BeautifulSoup(m["content"], "html.parser")
        table = soup.find("table")
        if not table:
            continue
        if headers is None:
            headers = [th.get_text(strip=True) for th in table.find_all("th")]
        for tr in table.find_all("tr"):
            cells = tr.find_all("td")
            if cells:
                rows.append([td.get_text(strip=True) for td in cells])
    return headers, rows


def find_col(headers, possible_names):
    for i, h in enumerate(headers):
        if h.strip() in possible_names:
            return i
    return None


# ---------------------------------------------------------------------------
# 3. TIỆN ÍCH: TRÍCH KHÓA TRÙNG LẶP (ưu tiên email, sau đó số điện thoại)
# ---------------------------------------------------------------------------
def extract_dedup_key(contact_text):
    if not contact_text:
        return None
    email_match = re.search(r"[\w\.\-]+@[\w\.\-]+\.\w+", contact_text)
    if email_match:
        return "email:" + email_match.group(0).strip().lower()
    phone_match = re.search(r"(0\d[\d\s\.]{7,})", contact_text)
    if phone_match:
        digits = re.sub(r"\D", "", phone_match.group(1))
        return "phone:" + digits
    return "raw:" + contact_text.strip().lower()


def parse_start_date(text):
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text or "")
    if match:
        try:
            return datetime.datetime.strptime(match.group(1), "%d/%m/%Y")
        except ValueError:
            pass
    return datetime.datetime.min


# ---------------------------------------------------------------------------
# 4. ĐỌC DỮ LIỆU ĐÃ CÓ TRONG FILE CŨ (nếu tồn tại)
# ---------------------------------------------------------------------------
def load_existing_rows(path, expected_headers):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    file_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if file_headers != expected_headers:
        print("Cảnh báo: cấu trúc cột trong file cũ khác với dữ liệu mới lấy về. "
              "Vẫn cố gắng gộp theo đúng thứ tự cột hiện có.")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in r):
            rows.append([("" if v is None else str(v)) for v in r])
    return rows


# ---------------------------------------------------------------------------
# 5. GỘP DỮ LIỆU CŨ + MỚI, KHỬ TRÙNG THEO SĐT/EMAIL, SẮP XẾP
# ---------------------------------------------------------------------------
def merge_and_sort(headers, all_rows):
    contact_col = find_col(headers, ["Số điện thoại - Email", "SĐT - Email", "Số điện thoại-Email"])
    start_date_col = find_col(headers, ["Ngày bắt đầu làm việc", "Ngày bắt đầu"])
    stt_col = find_col(headers, ["STT"])

    if contact_col is None or start_date_col is None:
        raise RuntimeError(f"Không xác định được cột SĐT/Email hoặc Ngày bắt đầu. Header đọc được: {headers}")

    merged = {}
    for row in all_rows:
        key = extract_dedup_key(row[contact_col])
        if key is None:
            continue
        start_dt = parse_start_date(row[start_date_col])
        if key not in merged or start_dt > parse_start_date(merged[key][start_date_col]):
            merged[key] = row

    final_rows = sorted(merged.values(), key=lambda r: parse_start_date(r[start_date_col]))
    return final_rows, stt_col


# ---------------------------------------------------------------------------
# 6. GHI FILE EXCEL (GHI ĐÈ LÊN CÙNG 1 FILE MỖI LẦN CHẠY)
# ---------------------------------------------------------------------------
def write_excel(path, headers, final_rows, stt_col, start_date_col_idx=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhân sự mới"

    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    body_font = Font(name=FONT_NAME, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    numeric_align_cols = {stt_col, start_date_col_idx}
    for i, row in enumerate(final_rows, start=2):
        for j, val in enumerate(row, start=1):
            v = (i - 1) if (stt_col is not None and j - 1 == stt_col) else val
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if (j - 1) in numeric_align_cols else left

    widths = [8] + [22] * (len(headers) - 1)
    if stt_col is not None:
        widths[stt_col] = 6
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths[idx - 1]

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    messages = fetch_messages()
    print(f"Đã lấy {len(messages)} tin nhắn từ Zulip.")

    headers, new_rows = parse_messages(messages)
    if headers is None:
        raise RuntimeError("Không tìm thấy bảng dữ liệu nào trong các tin nhắn của topic này.")
    print(f"Đã trích xuất {len(new_rows)} dòng nhân sự từ các tin nhắn.")

    existing_rows = load_existing_rows(OUTPUT_FILE, headers)
    if existing_rows:
        print(f"Đã đọc {len(existing_rows)} dòng có sẵn trong file {OUTPUT_FILE}.")

    all_rows = existing_rows + new_rows
    final_rows, stt_col = merge_and_sort(headers, all_rows)
    start_date_col_idx = find_col(headers, ["Ngày bắt đầu làm việc", "Ngày bắt đầu"])

    print(f"Sau khi khử trùng theo SĐT/Email: còn {len(final_rows)} nhân sự.")

    write_excel(OUTPUT_FILE, headers, final_rows, stt_col, start_date_col_idx)
    print(f"Đã cập nhật file: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
